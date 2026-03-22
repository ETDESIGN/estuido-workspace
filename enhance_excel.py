#!/usr/bin/env python3
"""
Enhance Excel with advanced formatting: borders, zebra striping, currency format
"""

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

# Load the generated Excel file
wb = load_workbook('Flash360_Quotation_Modern.xlsx')
ws = wb.active

# Define border styles
thick_border = Border(
    left=Side(style='thick', color='1F4E78'),
    right=Side(style='thick', color='1F4E78'),
    top=Side(style='thick', color='1F4E78'),
    bottom=Side(style='thick', color='1F4E78')
)

thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

# Define zebra stripe fill
zebra_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

# Apply formatting to the entire used range
max_row = ws.max_row
max_col = ws.max_column

print(f"Applying formatting to {max_row} rows x {max_col} columns...")

# Add thick outer border around the entire quote (from row 1 to max_row)
for row in range(1, max_row + 1):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        # Add thin border to all cells
        cell.border = thin_border

# Add thick outer border
for col in range(1, max_col + 1):
    # Top border
    ws.cell(row=1, column=col).border = Border(
        top=Side(style='thick', color='1F4E78'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    # Bottom border
    ws.cell(row=max_row, column=col).border = Border(
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thick', color='1F4E78')
    )

for row in range(1, max_row + 1):
    # Left border
    ws.cell(row=row, column=1).border = Border(
        left=Side(style='thick', color='1F4E78'),
        right=Side(style='thin', color='D0D0D0')
    )
    # Right border
    ws.cell(row=row, column=max_col).border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thick', color='1F4E78')
    )

# Apply zebra striping to data rows (rows 8-16 are the item rows)
for row in range(8, 17):
    if row % 2 == 0:  # Even rows get zebra stripe
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                cell.fill = zebra_fill

# Apply currency formatting to price/amount columns (columns D and E, starting from row 8)
for row in range(8, 21):
    for col in [4, 5]:  # Columns D and E
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '€#,##0.00'  # Euro currency format

# Set column widths for better auto-layout
ws.column_dimensions['A'].width = 8   # Item #
ws.column_dimensions['B'].width = 50  # Description (wide for full text)
ws.column_dimensions['C'].width = 12  # Quantity
ws.column_dimensions['D'].width = 15  # Unit Price
ws.column_dimensions['E'].width = 15  # Amount

# Set row heights for better readability
for row in range(1, max_row + 1):
    if row in [9, 11, 14, 16]:  # Detail rows with indented descriptions
        ws.row_dimensions[row].height = 30
    else:
        ws.row_dimensions[row].height = 20

# Add a thick border around the title section
for col in range(1, max_col + 1):
    cell = ws.cell(row=1, column=col)
    cell.border = thick_border

# Save the enhanced file
wb.save('Flash360_Quotation_Modern.xlsx')
print("✅ Excel file enhanced with borders, zebra striping, and currency formatting!")
print("📄 File saved: Flash360_Quotation_Modern.xlsx")
