#!/usr/bin/env python3
"""
Add clean, minimalist formatting to quotation
"""

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Load the generated Excel file
wb = load_workbook('Flash360_Quotation_Minimal.xlsx')
ws = wb.active

# Define border styles
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

thick_border = Border(
    left=Side(style='medium', color='000000'),
    right=Side(style='medium', color='000000'),
    top=Side(style='medium', color='000000'),
    bottom=Side(style='medium', color='000000')
)

# Light gray fill for header
header_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')

max_row = ws.max_row
max_col = ws.max_column

print(f"Formatting {max_row} rows x {max_col} columns...")

# 1. Format title rows (1-2) - centered, larger font
for row in [1, 2]:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if row == 1:
            cell.font = Font(size=16, bold=True)
        else:
            cell.font = Font(size=12)

# 2. Format quotation info (rows 4-5) - clean alignment
for row in [4, 5]:
    for col in [1, 2, 4, 5]:
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(vertical='center')

# 3. Format table header (row 7) - gray background, borders, centered
for col in range(1, max_col + 1):
    cell = ws.cell(row=7, column=col)
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, size=11)

# 4. Format data rows (8-11) - borders, proper alignment
for row in range(8, 12):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

        # Center the item numbers
        if col == 1:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# 5. Format summary rows (13-15) - borders, right-aligned labels
for row in [13, 14, 15]:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

        # Right-align labels and values in columns D and E
        if col in [4, 5]:
            cell.alignment = Alignment(horizontal='right', vertical='center')

# Make total row bold
for col in range(1, max_col + 1):
    cell = ws.cell(row=15, column=col)
    cell.font = Font(bold=True)

# 6. Add thick border around the entire table (rows 7-15, columns A-E)
# Top border
for col in range(1, 6):
    cell = ws.cell(row=7, column=col)
    cell.border = Border(
        top=Side(style='medium', color='000000'),
        bottom=Side(style='thin', color='000000'),
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000')
    )

# Bottom border
for col in range(1, 6):
    cell = ws.cell(row=15, column=col)
    cell.border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='medium', color='000000'),
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000')
    )

# Left and right borders
for row in range(7, 16):
    # Left border
    cell = ws.cell(row=row, column=1)
    cell.border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
        left=Side(style='medium', color='000000'),
        right=Side(style='thin', color='000000')
    )
    # Right border
    cell = ws.cell(row=row, column=5)
    cell.border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
        left=Side(style='thin', color='000000'),
        right=Side(style='medium', color='000000')
    )

# 7. Format notes section (rows 17-20) - clean, no borders
for row in range(17, 21):
    cell = ws.cell(row=row, column=1)
    cell.alignment = Alignment(vertical='top', wrap_text=True)
    ws.row_dimensions[row].height = 20

# 8. Format terms section (rows 22-23)
cell = ws.cell(row=23, column=1)
cell.alignment = Alignment(vertical='top', wrap_text=True)
ws.row_dimensions[23].height = 30

# 9. Format signature section (rows 25-27)
for row in range(25, 28):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(vertical='center')

# 10. Format footer (row 29)
for col in range(1, max_col + 1):
    cell = ws.cell(row=29, column=col)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(italic=True)

# 11. Set column widths for clean layout
ws.column_dimensions['A'].width = 8     # Item
ws.column_dimensions['B'].width = 40    # Description
ws.column_dimensions['C'].width = 12    # Quantity
ws.column_dimensions['D'].width = 15    # Unit Price
ws.column_dimensions['E'].width = 15    # Amount

# 12. Set row heights for better spacing
ws.row_dimensions[1].height = 25        # Title
ws.row_dimensions[2].height = 20        # Subtitle
for row in range(4, 6):
    ws.row_dimensions[row].height = 20  # Info rows
for row in range(7, 12):
    ws.row_dimensions[row].height = 25  # Data rows
for row in range(13, 16):
    ws.row_dimensions[row].height = 22  # Summary rows

# 13. Apply currency format to price columns (D and E, rows 8-15)
for row in range(8, 16):
    for col in [4, 5]:
        cell = ws.cell(row=row, column=col)
        cell.number_format = '€#,##0.00'

# Save the formatted file
wb.save('Flash360_Quotation_Minimal.xlsx')
print("✅ Minimalist Excel quotation created successfully!")
print("📄 Features: Clean borders, proper centering, no extra colors")
