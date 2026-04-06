#!/usr/bin/env python3
"""Generate white-label quotation: Excel + PDF for TJ0066"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_LEFT, TA_CENTER
import os

OUTPUT_DIR = "/home/e/Desktop"

# ─── Styles ───
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
bold = Font(name='Arial', bold=True, size=11)
bold_header = Font(name='Arial', bold=True, size=14)
bold_sub = Font(name='Arial', bold=True, size=12)
normal = Font(name='Arial', size=11)
eur_format = '#,##0.00 "€"'

def style_cell(cell, font=normal, border=thin_border, align=None, wrap=True):
    cell.font = font
    cell.border = border
    if align:
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    else:
        cell.alignment = Alignment(vertical='center', wrap_text=wrap)

# ═══════════════════════════════════════════
# EXCEL WORKBOOK
# ═══════════════════════════════════════════
wb = Workbook()

# ─── Sheet 1: RFQ to Supplier ───
ws1 = wb.active
ws1.title = "RFQ to Supplier"
ws1.sheet_properties.tabColor = "999999"

# Column widths
col_widths = [6, 30, 12, 26, 26, 18, 40]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title
ws1.merge_cells('A1:G1')
c = ws1['A1']
c.value = "RFQ to Supplier"
c.font = bold_sub
c.alignment = Alignment(horizontal='left', vertical='center')

# Header block
ws1.merge_cells('A3:G3')
c = ws1['A3']
c.value = "Request for Quotation — TJ0066"
c.font = bold_header
c.alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('A4:G4')
c = ws1['A4']
c.value = "Currency: EUR   |   Incoterm: EXW"
c.font = bold
c.alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('A5:G5')
c = ws1['A5']
c.value = "Dear Supplier, please provide your BEST unit prices (EUR, EXW) for the items below and confirm the technical points listed."
c.font = normal
c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Table headers (row 7)
headers = ["Item", "Model/Spec", "Quantity", "Target selling\nprice (EUR)", 
           "Supplier unit\nprice (EUR)", "Line total\n(EUR)", "Notes"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=7, column=col, value=h)
    style_cell(c, font=bold, align='center')

# Data rows
rows_data = [
    [1, "Dock — 8-battery", "TJ0066 (8 slots)", 150, 76.80, None, "Please quote your BEST price."],
    [2, "Dock — 12-battery", "TJ0066 (12 slots)", 80, None, None, "Please quote your BEST price."],
    [3, "Power bank — 5000 mAh", "BST-0146 spec", 500, 9.85, None, "Please quote your BEST price for 500 pcs and 2000 pcs (two separate rows)."],
    [4, "Power bank — 5000 mAh", "BST-0146 spec", 2000, 9.85, None, ""],
    [5, "Dock — 4-battery", "BST-0264 (4 slots, no screen)", 20, None, None, "Previous quoted price for 4-bay is too high; please re-quote your BEST price."],
]

for i, row in enumerate(rows_data):
    r = 8 + i
    for col, val in enumerate(row, 1):
        c = ws1.cell(row=r, column=col, value=val)
        if col == 1:
            style_cell(c, align='center')
        elif col == 3:
            style_cell(c, align='center')
            c.number_format = '#,##0'
        elif col in (4, 5):
            style_cell(c, align='right')
            if val is not None:
                c.number_format = eur_format
        elif col == 6:
            # Line total formula: Supplier unit price * Quantity
            formula = f'=IF(E{r}="","",E{r}*C{r})'
            c = ws1.cell(row=r, column=col, value=formula)
            style_cell(c, align='right')
            c.number_format = eur_format
        elif col == 7:
            style_cell(c, align='left')
        else:
            style_cell(c, align='left')

# Total row
total_row = 13
ws1.merge_cells(f'A{total_row}:E{total_row}')
c = ws1.cell(row=total_row, column=1, value="Total:")
style_cell(c, font=bold, align='right')
c = ws1.cell(row=total_row, column=6, value=f'=SUM(F8:F12)')
style_cell(c, font=bold, align='right')
c.number_format = eur_format
for col in range(7, 8):
    style_cell(ws1.cell(row=total_row, column=col))

# Spacer
spacer_row = 15

# Technical checklist
ws1.merge_cells(f'A{spacer_row}:G{spacer_row}')
c = ws1.cell(row=spacer_row, column=1, value="Technical Checklist:")
c.font = bold

checklist = [
    "• Pogo-pin charging only on battery (no USB port)",
    "• Charging at 5V / 1A (no fast charge)",
    "• Each battery with unique serial number (SN) readable via UART or equivalent",
    "• Data available to computer: SN, battery level, cycle count, presence, release",
    "• Sound feedback when battery inserted",
    "• Reinforced pogo-pin mount; metal latches; proper cable channels; spring mechanism",
    "• Screen printing on battery (1–2 colors) — confirm feasibility and cost",
]
for j, item in enumerate(checklist):
    r = spacer_row + 1 + j
    ws1.merge_cells(f'A{r}:G{r}')
    c = ws1.cell(row=r, column=1, value=item)
    c.font = normal

# Closing
close_row = spacer_row + len(checklist) + 2
ws1.merge_cells(f'A{close_row}:G{close_row}')
c = ws1.cell(row=close_row, column=1, value="Please send your BEST unit prices (EUR, EXW) for the items above. Thank you.")
c.font = normal

sig_row = close_row + 2
ws1.merge_cells(f'A{sig_row}:G{sig_row}')
c = ws1.cell(row=sig_row, column=1, value="Best regards,")
c.font = normal

sig_row2 = close_row + 3
ws1.merge_cells(f'A{sig_row2}:G{sig_row2}')
c = ws1.cell(row=sig_row2, column=1, value="")

# Print settings
ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
ws1.page_setup.orientation = 'portrait'
ws1.page_margins.left = 15*mm
ws1.page_margins.right = 15*mm
ws1.page_margins.top = 15*mm
ws1.page_margins.bottom = 15*mm


# ─── Sheet 2: Product Specification ───
ws2 = wb.create_sheet("Product Specification")
ws2.sheet_properties.tabColor = "999999"

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 55

# Title
ws2.merge_cells('A1:B1')
c = ws2['A1']
c.value = "Product Specification"
c.font = bold_sub
c.alignment = Alignment(horizontal='left', vertical='center')

# Spec sections
specs = {
    "Consumption": [
        ("Electricity (Standby/Month)", "8 KWH"),
        ("Roaming Data Usage (Month)", "20 MB"),
    ],
    "Security": [
        ("Material", "V0 Flame Resistance ABS"),
        ("Overvoltage Protection", "Yes"),
        ("Protection from Station", "Rain-resistant (mechanical)"),
        ("Drop Test", "1.2M drop test pass"),
        ("Certification", "CB certified adapter and module"),
    ],
    "Hardware": [
        ("Model Name", "TJ0066"),
        ("Slot Option", "8 slots"),
        ("Communication", "4G (WIFI optional)"),
        ("Dual SIM", "Network connection enhanced"),
        ("Battery Option", "5000mAh / 8000mAh / 10000mAh"),
    ],
    "Specification": [
        ("AC Input", "AC100V–240V"),
        ("DC Output", "5V 8A"),
        ("Each Slot Output", "10W"),
        ("Operating Temperature", "-10°C ~ 45°C"),
        ("Operating Humidity", "5% ~ 95%"),
        ("Product Size", "L190 × W240 × H280 mm"),
        ("Packaging Size", "30 × 28 × 41 cm"),
        ("Weight (Without Power Banks)", "2.35 kg"),
        ("Color", "Customizable"),
        ("Certificates", "CE / CB / RoHS / MSDS / UN38.3"),
    ],
}

r = 3
for section, items in specs.items():
    # Section header
    ws2.merge_cells(f'A{r}:B{r}')
    c = ws2.cell(row=r, column=1, value=section)
    style_cell(c, font=bold)
    c.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    style_cell(ws2.cell(row=r, column=2))
    ws2.cell(row=r, column=2).fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    r += 1
    
    # Table header
    c1 = ws2.cell(row=r, column=1, value="Parameter")
    c2 = ws2.cell(row=r, column=2, value="Value")
    style_cell(c1, font=bold, align='left')
    style_cell(c2, font=bold, align='left')
    c1.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    c2.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    r += 1
    
    for param, val in items:
        c1 = ws2.cell(row=r, column=1, value=param)
        c2 = ws2.cell(row=r, column=2, value=val)
        style_cell(c1, align='left')
        style_cell(c2, align='left')
        r += 1
    
    r += 1  # spacer between sections

ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
ws2.page_setup.orientation = 'portrait'
ws2.page_margins.left = 15*mm
ws2.page_margins.right = 15*mm
ws2.page_margins.top = 15*mm
ws2.page_margins.bottom = 15*mm

# Save Excel
excel_path = os.path.join(OUTPUT_DIR, "Quotation_TJ0066_2026_whitelabel.xlsx")
wb.save(excel_path)
print(f"✅ Excel saved: {excel_path}")


# ═══════════════════════════════════════════
# PDF GENERATION
# ═══════════════════════════════════════════
pdf_path = os.path.join(OUTPUT_DIR, "Quotation_TJ0066_2026_whitelabel.pdf")

doc = SimpleDocTemplate(
    pdf_path,
    pagesize=A4,
    leftMargin=15*mm,
    rightMargin=15*mm,
    topMargin=15*mm,
    bottomMargin=15*mm,
)

styles = getSampleStyleSheet()
style_title = ParagraphStyle('Title', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=14, spaceAfter=6)
style_h1 = ParagraphStyle('H1', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=13, spaceAfter=8, spaceBefore=12, alignment=TA_CENTER)
style_h2 = ParagraphStyle('H2', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11, spaceAfter=6)
style_body = ParagraphStyle('Body', parent=styles['Normal'], fontName='Helvetica', fontSize=11, spaceAfter=6, leading=14)
style_body_sm = ParagraphStyle('BodySm', parent=styles['Normal'], fontName='Helvetica', fontSize=10, spaceAfter=3, leading=13)
style_section = ParagraphStyle('Section', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11, spaceAfter=4, spaceBefore=8)

# Colors
gray_border = colors.Color(0.85, 0.85, 0.85)
light_fill = colors.Color(0.95, 0.95, 0.95)
header_fill = colors.Color(0.90, 0.90, 0.90)

story = []

# ─── PAGE 1: RFQ ───
story.append(Paragraph("RFQ to Supplier", style_title))
story.append(Spacer(1, 4))
story.append(Paragraph("Request for Quotation — TJ0066", style_h1))
story.append(Paragraph("Currency: EUR  |  Incoterm: EXW", ParagraphStyle('Center', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11, alignment=TA_CENTER, spaceAfter=10)))
story.append(Paragraph("Dear Supplier, please provide your BEST unit prices (EUR, EXW) for the items below and confirm the technical points listed.", style_body))
story.append(Spacer(1, 6))

# RFQ Table
table_headers = ["Item", "Model/Spec", "Qty", "Target price\n(EUR)", "Supplier price\n(EUR)", "Line total\n(EUR)", "Notes"]
table_data = [
    table_headers,
    ["1", "Dock — 8-battery", "150", "€ 76.80", "", "", "Please quote your BEST price."],
    ["2", "Dock — 12-battery", "80", "", "", "", "Please quote your BEST price."],
    ["3", "Power bank — 5000 mAh", "500", "€ 9.85", "", "", "Please quote for 500 pcs and 2000 pcs (two rows)."],
    ["4", "Power bank — 5000 mAh", "2,000", "€ 9.85", "", "", ""],
    ["5", "Dock — 4-battery", "20", "", "", "", "Previous 4-bay price too high; please re-quote."],
    ["", "", "", "", "Total:", "", ""],
]

col_widths = [25, 95, 30, 55, 55, 50, 150]
t = Table(table_data, colWidths=col_widths, repeatRows=1)
t.setStyle(TableStyle([
    # Header
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('FONTSIZE', (0, 0), (-1, 0), 9),
    ('BACKGROUND', (0, 0), (-1, 0), header_fill),
    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    # Body
    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
    ('FONTSIZE', (0, 1), (-1, -1), 9),
    ('ALIGN', (0, 1), (0, -1), 'CENTER'),
    ('ALIGN', (2, 1), (2, -1), 'CENTER'),
    ('ALIGN', (3, 1), (5, -1), 'RIGHT'),
    # Borders
    ('GRID', (0, 0), (-1, -1), 0.5, gray_border),
    ('BOX', (0, 0), (-1, -1), 1, colors.Color(0.6, 0.6, 0.6)),
    # Total row
    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ('LINEABOVE', (0, -1), (-1, -1), 1, colors.Color(0.4, 0.4, 0.4)),
]))
story.append(t)
story.append(Spacer(1, 12))

# Technical checklist
story.append(Paragraph("Technical Checklist", style_h2))
checklist_items = [
    "• Pogo-pin charging only on battery (no USB port)",
    "• Charging at 5V / 1A (no fast charge)",
    "• Each battery with unique serial number (SN) readable via UART or equivalent",
    "• Data available to computer: SN, battery level, cycle count, presence, release",
    "• Sound feedback when battery inserted",
    "• Reinforced pogo-pin mount; metal latches; proper cable channels; spring mechanism",
    "• Screen printing on battery (1–2 colors) — confirm feasibility and cost",
]
for item in checklist_items:
    story.append(Paragraph(item, style_body_sm))

story.append(Spacer(1, 14))
story.append(Paragraph("Please send your BEST unit prices (EUR, EXW) for the items above. Thank you.", style_body))
story.append(Spacer(1, 20))
story.append(Paragraph("Best regards,", style_body))

# ─── PAGE BREAK → PAGE 2: Specification ───
story.append(PageBreak())

story.append(Paragraph("Product Specification", style_title))
story.append(Spacer(1, 10))

for section, items in specs.items():
    story.append(Paragraph(section, style_section))
    
    sec_data = [["Parameter", "Value"]]
    for param, val in items:
        sec_data.append([param, val])
    
    sec_t = Table(sec_data, colWidths=[170, 290])
    sec_t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), header_fill),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, gray_border),
        ('BOX', (0, 0), (-1, -1), 1, colors.Color(0.6, 0.6, 0.6)),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(sec_t)
    story.append(Spacer(1, 6))

# Build PDF
doc.build(story)
print(f"✅ PDF saved: {pdf_path}")

print("\n✅ Both files generated successfully!")
