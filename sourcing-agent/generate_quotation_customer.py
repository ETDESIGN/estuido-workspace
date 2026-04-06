#!/usr/bin/env python3
"""Generate white-label quotation: Excel + PDF for TJ0066 (CUSTOMER-FACING)
This is a quotation FROM E-Studio TO a customer, showing selling prices.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

OUTPUT_DIR = "/home/e/Desktop"

thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
bold = Font(name='Arial', bold=True, size=11)
bold_header = Font(name='Arial', bold=True, size=14)
bold_sub = Font(name='Arial', bold=True, size=12)
normal = Font(name='Arial', size=11)
eur_format = '#,##0.00 "€"'

def style_cell(cell, font=normal, border=thin_border, align=None, wrap=True):
    cell.font = font
    cell.border = border
    if align:
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    else:
        cell.alignment = Alignment(vertical='center', wrap_text=wrap)

# ═══════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════
wb = Workbook()

# ─── Sheet 1: Quotation ───
ws1 = wb.active
ws1.title = "Quotation"
ws1.sheet_properties.tabColor = "999999"

col_widths = [6, 30, 14, 20, 20, 18, 35]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title
ws1.merge_cells('A1:G1')
ws1['A1'].value = "Quotation"
ws1['A1'].font = bold_sub

# Header block
ws1.merge_cells('A3:G3')
ws1['A3'].value = "Quotation — TJ0066"
ws1['A3'].font = bold_header
ws1['A3'].alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('A4:G4')
ws1['A4'].value = "Currency: EUR   |   Incoterm: EXW"
ws1['A4'].font = bold
ws1['A4'].alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('A5:G5')
ws1['A5'].value = "Date: March 31, 2026   |   Valid for 30 days"
ws1['A5'].font = normal
ws1['A5'].alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('A7:G7')
ws1['A7'].value = "Dear Customer, please find below our quotation for the items discussed. All prices are in EUR, EXW."
ws1['A7'].font = normal
ws1['A7'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Table headers (row 9)
headers = ["Item", "Description", "Quantity", "Unit Price\n(EUR, EXW)", "Line Total\n(EUR)", "Lead\nTime", "Notes"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=9, column=col, value=h)
    style_cell(c, font=bold, align='center')

# Data rows — CUSTOMER-FACING PRICES
rows_data = [
    [1, "Dock — 8-battery charging station\nTJ0066 (8 slots)", 150, 76.80, None, "4–6 weeks", "Includes 8 batteries per dock"],
    [2, "Dock — 12-battery charging station\nTJ0066 (12 slots)", 80, 99.50, None, "4–6 weeks", "Includes 12 batteries per dock"],
    [3, "Power bank — 5000 mAh\nBST-0146 spec", 500, 9.85, None, "2–3 weeks", "Pogo-pin charging only, no USB port"],
    [4, "Power bank — 5000 mAh\nBST-0146 spec (bulk)", 2000, 8.50, None, "3–4 weeks", "Volume discount applied"],
    [5, "Dock — 4-battery charging station\nBST-0264 (4 slots, no screen)", 20, 54.00, None, "4–6 weeks", "Compact model, no LCD screen"],
]

for i, row in enumerate(rows_data):
    r = 10 + i
    for col, val in enumerate(row, 1):
        c = ws1.cell(row=r, column=col, value=val)
        if col == 1:
            style_cell(c, align='center')
        elif col == 3:
            style_cell(c, align='center')
            c.number_format = '#,##0'
        elif col in (4, 5):
            style_cell(c, align='right')
            if val is not None:
                c.number_format = eur_format
        elif col == 6:
            # Line total formula: Unit Price * Quantity
            formula = f'=IF(D{r}="","",D{r}*C{r})'
            c = ws1.cell(row=r, column=6, value=formula)
            style_cell(c, align='right')
            c.number_format = eur_format
        elif col == 7:
            style_cell(c, align='left')
        else:
            style_cell(c, align='left')

# Total row
total_row = 15
ws1.merge_cells(f'A{total_row}:D{total_row}')
c = ws1.cell(row=total_row, column=1, value="Total:")
style_cell(c, font=bold, align='right')
c = ws1.cell(row=total_row, column=6, value=f'=SUM(F10:F14)')
style_cell(c, font=bold, align='right')
c.number_format = eur_format

# Terms
terms_start = 17
ws1.merge_cells(f'A{terms_start}:G{terms_start}')
ws1[f'A{terms_start}'].value = "Terms & Conditions"
ws1[f'A{terms_start}'].font = bold_sub

terms = [
    "• Prices are in EUR, EXW (Ex Works) Dongguan, China",
    "• Quotation valid for 30 days from date of issue",
    "• Payment: 30% deposit with order, 70% before shipment",
    "• Tooling costs are included in unit price for orders above MOQ",
    "• Lead times start from confirmation of order and deposit receipt",
    "• Certifications: CE, CB, RoHS, MSDS, UN38.3 (included)",
    "• Shipping and customs clearance: buyer's responsibility",
    "• All dimensions and specifications subject to ±2% manufacturing tolerance",
]
for j, term in enumerate(terms):
    r = terms_start + 1 + j
    ws1.merge_cells(f'A{r}:G{r}')
    ws1[f'A{r}'].value = term
    ws1[f'A{r}'].font = normal

# Closing
close_row = terms_start + len(terms) + 2
ws1.merge_cells(f'A{close_row}:G{close_row}')
ws1[f'A{close_row}'].value = "Should you have any questions, please do not hesitate to contact us."
ws1[f'A{close_row}'].font = normal

ws1.merge_cells(f'A{close_row+2}:G{close_row+2}')
ws1[f'A{close_row+2}'].value = "Best regards,"
ws1[f'A{close_row+2}'].font = normal

ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
ws1.page_setup.orientation = 'portrait'
ws1.page_margins.left = 15*mm
ws1.page_margins.right = 15*mm
ws1.page_margins.top = 15*mm
ws1.page_margins.bottom = 15*mm


# ─── Sheet 2: Product Specification ───
ws2 = wb.create_sheet("Product Specification")
ws2.sheet_properties.tabColor = "999999"
ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 55

ws2.merge_cells('A1:B1')
ws2['A1'].value = "Product Specification"
ws2['A1'].font = bold_sub

specs = {
    "Consumption": [
        ("Electricity (Standby/Month)", "8 KWH"),
        ("Roaming Data Usage (Month)", "20 MB"),
    ],
    "Security": [
        ("Material", "V0 Flame Resistance ABS"),
        ("Overvoltage Protection", "Yes"),
        ("Protection from Station", "Rain-resistant (mechanical)"),
        ("Drop Test", "1.2M drop test pass"),
        ("Certification", "CB certified adapter and module"),
    ],
    "Hardware": [
        ("Model Name", "TJ0066"),
        ("Slot Option", "8 slots"),
        ("Communication", "4G (WIFI optional)"),
        ("Dual SIM", "Network connection enhanced"),
        ("Battery Option", "5000mAh / 8000mAh / 10000mAh"),
    ],
    "Specification": [
        ("AC Input", "AC100V–240V"),
        ("DC Output", "5V 8A"),
        ("Each Slot Output", "10W"),
        ("Operating Temperature", "-10°C ~ 45°C"),
        ("Operating Humidity", "5% ~ 95%"),
        ("Product Size", "L190 × W240 × H280 mm"),
        ("Packaging Size", "30 × 28 × 41 cm"),
        ("Weight (Without Power Banks)", "2.35 kg"),
        ("Color", "Customizable"),
        ("Certificates", "CE / CB / RoHS / MSDS / UN38.3"),
    ],
}

r = 3
for section, items in specs.items():
    ws2.merge_cells(f'A{r}:B{r}')
    ws2[f'A{r}'].value = section
    style_cell(ws2[f'A{r}'], font=bold)
    ws2[f'A{r}'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    style_cell(ws2[f'B{r}'])
    ws2[f'B{r}'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    r += 1
    for param, val in items:
        c1 = ws2.cell(row=r, column=1, value=param)
        c2 = ws2.cell(row=r, column=2, value=val)
        style_cell(c1, align='left')
        style_cell(c2, align='left')
        r += 1
    r += 1

ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
ws2.page_setup.orientation = 'portrait'
ws2.page_margins.left = 15*mm
ws2.page_margins.right = 15*mm
ws2.page_margins.top = 15*mm
ws2.page_margins.bottom = 15*mm

excel_path = f"{OUTPUT_DIR}/Quotation_TJ0066_2026_whitelabel_CUSTOMER.xlsx"
wb.save(excel_path)
print(f"✅ Excel: {excel_path}")


# ═══════════════════════════════════════════
# PDF
# ═══════════════════════════════════════════
pdf_path = f"{OUTPUT_DIR}/Quotation_TJ0066_2026_whitelabel_CUSTOMER.pdf"

doc = SimpleDocTemplate(pdf_path, pagesize=A4,
    leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)

gray_border = colors.Color(0.85, 0.85, 0.85)
header_fill = colors.Color(0.90, 0.90, 0.90)

style_h1 = ParagraphStyle('H1', fontName='Helvetica-Bold', fontSize=13, alignment=TA_CENTER, spaceAfter=4)
style_h2 = ParagraphStyle('H2', fontName='Helvetica-Bold', fontSize=11, spaceAfter=4)
style_body = ParagraphStyle('Body', fontName='Helvetica', fontSize=11, spaceAfter=6, leading=14)
style_sm = ParagraphStyle('Sm', fontName='Helvetica', fontSize=10, spaceAfter=3, leading=13)
style_section = ParagraphStyle('Sec', fontName='Helvetica-Bold', fontSize=11, spaceAfter=4, spaceBefore=8)

story = []

# ─── PAGE 1: Quotation ───
story.append(Paragraph("Quotation", ParagraphStyle('Title', fontName='Helvetica-Bold', fontSize=14, spaceAfter=6)))
story.append(Spacer(1, 4))
story.append(Paragraph("Quotation — TJ0066", style_h1))
story.append(Paragraph("Currency: EUR  |  Incoterm: EXW", ParagraphStyle('Center', fontName='Helvetica-Bold', fontSize=11, alignment=TA_CENTER, spaceAfter=4)))
story.append(Paragraph("Date: March 31, 2026  |  Valid for 30 days", ParagraphStyle('Center', fontName='Helvetica', fontSize=11, alignment=TA_CENTER, spaceAfter=10)))
story.append(Paragraph("Dear Customer, please find below our quotation for the items discussed. All prices are in EUR, EXW.", style_body))
story.append(Spacer(1, 6))

# Quotation table
table_headers = ["Item", "Description", "Qty", "Unit Price\n(EUR, EXW)", "Line Total\n(EUR)", "Lead Time", "Notes"]
table_data = [
    table_headers,
    ["1", "Dock — 8-battery charging station (TJ0066)", "150", "€ 76.80", "€ 11,520.00", "4–6 weeks", "Includes 8 batteries"],
    ["2", "Dock — 12-battery charging station (TJ0066)", "80", "€ 99.50", "€ 7,960.00", "4–6 weeks", "Includes 12 batteries"],
    ["3", "Power bank — 5000 mAh (BST-0146)", "500", "€ 9.85", "€ 4,925.00", "2–3 weeks", "Pogo-pin charging"],
    ["4", "Power bank — 5000 mAh (bulk)", "2,000", "€ 8.50", "€ 17,000.00", "3–4 weeks", "Volume discount"],
    ["5", "Dock — 4-battery station (BST-0264)", "20", "€ 54.00", "€ 1,080.00", "4–6 weeks", "No LCD screen"],
    ["", "", "", "", "Total:", "€ 42,485.00", ""],
]

col_widths = [25, 120, 30, 55, 55, 50, 95]
t = Table(table_data, colWidths=col_widths, repeatRows=1)
t.setStyle(TableStyle([
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('FONTSIZE', (0, 0), (-1, -1), 9),
    ('BACKGROUND', (0, 0), (-1, 0), header_fill),
    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
    ('ALIGN', (0, 1), (0, -1), 'CENTER'),
    ('ALIGN', (2, 1), (2, -1), 'CENTER'),
    ('ALIGN', (3, 1), (5, -1), 'RIGHT'),
    ('GRID', (0, 0), (-1, -1), 0.5, gray_border),
    ('BOX', (0, 0), (-1, -1), 1, colors.Color(0.6, 0.6, 0.6)),
    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ('LINEABOVE', (0, -1), (-1, -1), 1, colors.Color(0.4, 0.4, 0.4)),
    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ('TOPPADDING', (0, 0), (-1, -1), 4),
    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
]))
story.append(t)
story.append(Spacer(1, 12))

# Terms
story.append(Paragraph("Terms & Conditions", style_h2))
terms_text = [
    "• Prices are in EUR, EXW (Ex Works) Dongguan, China",
    "• Quotation valid for 30 days from date of issue",
    "• Payment: 30% deposit with order, 70% before shipment",
    "• Tooling costs included in unit price for orders above MOQ",
    "• Lead times start from confirmation of order and deposit receipt",
    "• Certifications: CE, CB, RoHS, MSDS, UN38.3 (included)",
    "• Shipping and customs clearance: buyer's responsibility",
    "• All dimensions and specifications subject to ±2% manufacturing tolerance",
]
for term in terms_text:
    story.append(Paragraph(term, style_sm))

story.append(Spacer(1, 14))
story.append(Paragraph("Should you have any questions, please do not hesitate to contact us.", style_body))
story.append(Spacer(1, 20))
story.append(Paragraph("Best regards,", style_body))

# ─── PAGE 2: Specification ───
story.append(PageBreak())
story.append(Paragraph("Product Specification", ParagraphStyle('Title', fontName='Helvetica-Bold', fontSize=14, spaceAfter=10)))

for section, items in specs.items():
    story.append(Paragraph(section, style_section))
    sec_data = [["Parameter", "Value"]]
    for param, val in items:
        sec_data.append([param, val])
    sec_t = Table(sec_data, colWidths=[170, 290])
    sec_t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), header_fill),
        ('GRID', (0, 0), (-1, -1), 0.5, gray_border),
        ('BOX', (0, 0), (-1, -1), 1, colors.Color(0.6, 0.6, 0.6)),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(sec_t)
    story.append(Spacer(1, 6))

doc.build(story)
print(f"✅ PDF: {pdf_path}")
print("\n✅ Customer-facing quotation generated!")
